"""
BizMind AI — File Parser Module
Parses both CSV and Excel (.xlsx) files using pure Python (csv and openpyxl).
This avoids C-assembly compile errors with Pandas on Python 3.14.
"""

import io
import csv
import openpyxl

def parse_csv(content: bytes):
    """Parse CSV content using built-in csv module."""
    text = content.decode('utf-8', errors='ignore')
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], 0, []
    
    columns = list(rows[0].keys())
    return rows, len(rows), columns

def parse_excel(content: bytes):
    """Parse Excel content using openpyxl."""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    sheet = wb.active
    
    rows = []
    columns = []
    
    # Read headers
    for cell in sheet[1]:
        val = str(cell.value).strip() if cell.value is not None else ""
        columns.append(val)
        
    for r in range(2, sheet.max_row + 1):
        row_dict = {}
        has_val = False
        for c in range(1, len(columns) + 1):
            header = columns[c-1]
            if not header:
                continue
            val = sheet.cell(row=r, column=c).value
            if val is not None:
                has_val = True
            row_dict[header] = val
        if has_val:
            rows.append(row_dict)
            
    # Filter out empty string columns
    columns = [col for col in columns if col]
    return rows, len(rows), columns

def process_file_data(content: bytes, ext: str):
    """Common processing for stats generation."""
    if ext == "csv":
        rows, total_rows, columns = parse_csv(content)
    else:
        rows, total_rows, columns = parse_excel(content)

    # Detect numeric columns and compute basic stats
    numeric_fields = []
    sums = {}
    averages = {}

    if rows:
        # Check first 50 rows to detect numeric fields
        for col in columns:
            if not col:
                continue
            is_numeric = True
            sample_count = 0
            for r in rows[:50]:
                val = r.get(col)
                if val is not None and val != "":
                    sample_count += 1
                    try:
                        float(str(val).replace(',', ''))
                    except ValueError:
                        is_numeric = False
                        break
            if is_numeric and sample_count > 0:
                numeric_fields.append(col)

        # Compute sums and averages
        for col in numeric_fields:
            total = 0.0
            valid_count = 0
            for r in rows:
                val = r.get(col)
                if val is not None and val != "":
                    try:
                        total += float(str(val).replace(',', ''))
                        valid_count += 1
                    except ValueError:
                        pass
            sums[col] = round(total, 2)
            averages[col] = round(total / valid_count, 2) if valid_count > 0 else 0

    return {
        "columns": columns,
        "total_rows": total_rows,
        "numeric_fields": numeric_fields,
        "sums": sums,
        "averages": averages,
        "preview": rows[:10],
        "all_rows": rows[:500]  # Return up to 500 rows for frontend charts
    }
